from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill


def alignment_31():
    ws["C1"].fill = PatternFill("solid", start_color="5cb800")
    ws["B3"].fill = PatternFill("solid", start_color="248ae6")
    ws["C1"].alignment = Alignment(horizontal="center")

    for cell in ws['B']:  # for B column
        cell.alignment = Alignment(horizontal='center')

    for rows in ws.iter_cols(min_col=2, max_col=2, min_row=3, max_row=21):
        for cell in rows:
            cell.fill = PatternFill("solid", start_color="248ae6")

    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=3, max_col=33):
        for cell in rows:
            cell.fill = PatternFill("solid", start_color="ffba72")

    ws.column_dimensions["B"].width = 15
    wb.save("budget.xlsx")

def alignment_28():
    ws["C1"].fill = PatternFill("solid", start_color="5cb800")
    ws["B3"].fill = PatternFill("solid", start_color="248ae6")
    ws["C1"].alignment = Alignment(horizontal="center")

    for cell in ws['B']:  # for B column
        cell.alignment = Alignment(horizontal='center')

    for rows in ws.iter_cols(min_col=2, max_col=2, min_row=3, max_row=21):
        for cell in rows:
            cell.fill = PatternFill("solid", start_color="248ae6")

    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=3, max_col=30):
        for cell in rows:
            cell.fill = PatternFill("solid", start_color="ffba72")

    ws.column_dimensions["B"].width = 15
    wb.save("budget.xlsx")

def alignment_30():
    ws["C1"].fill = PatternFill("solid", start_color="5cb800")
    ws["B3"].fill = PatternFill("solid", start_color="248ae6")
    ws["C1"].alignment = Alignment(horizontal="center")

    for cell in ws['B']:  # for B column
        cell.alignment = Alignment(horizontal='center')

    for rows in ws.iter_cols(min_col=2, max_col=2, min_row=3, max_row=21):
        for cell in rows:
            cell.fill = PatternFill("solid", start_color="248ae6")

    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=3, max_col=32):
        for cell in rows:
            cell.fill = PatternFill("solid", start_color="ffba72")

    ws.column_dimensions["B"].width = 15
    wb.save("budget.xlsx")

wb = load_workbook("budget.xlsx")
ws = wb["JANUARY"]
alignment_31()

wb = load_workbook("budget.xlsx")
ws = wb["FEBRUARY"]
alignment_28()

wb = load_workbook("budget.xlsx")
ws = wb["MARCH"]
alignment_31()

wb = load_workbook("budget.xlsx")
ws = wb["APRIL"]
alignment_30()

wb = load_workbook("budget.xlsx")
ws = wb["MAY"]
alignment_31()

wb = load_workbook("budget.xlsx")
ws = wb["JUNE"]
alignment_30()

wb = load_workbook("budget.xlsx")
ws = wb["JULY"]
alignment_31()

wb = load_workbook("budget.xlsx")
ws = wb["AUGUST"]
alignment_31()

wb = load_workbook("budget.xlsx")
ws = wb["SEPTEMBER"]
alignment_30()

wb = load_workbook("budget.xlsx")
ws = wb["OCTOBER"]
alignment_31()

wb = load_workbook("budget.xlsx")
ws = wb["NOVEMBER"]
alignment_30()

wb = load_workbook("budget.xlsx")
ws = wb["DECEMBER"]
alignment_31()


'''
ws1.column_dimensions["B"].width = 13
ws1.column_dimensions["C"].width = 13
ws1.column_dimensions["D"].width = 8
'''
'''
    pink = "00FF00FF"
    green = "00008000"
    thin = Side(border_style="thin", color=pink)
    double = Side(border_style="double", color=green)

    for cell in ws['1:1']:  # for 1 row:
        cell.alignment = Alignment(horizontal='center')

    ws["B3"].border = Border(top=thin, left=double, right=double, bottom=thin)
    '''