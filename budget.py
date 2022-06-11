from openpyxl import Workbook, load_workbook
import openpyxl
import datetime

months = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER",
          "DECEMBER"]
expenses = ["BILLS", "CAR", "SHOPPING", "EATING OUT", "CLOTHES", "HOME SHOPPING", "EVENTS", "GIFTS", "AGD RTV",
            "ENTERTAINMENT", "HOLIDAYS", "REPAIRS", "SPORT", "ANIMALS", "MEDICINE", "INSURANCE", "BEAUTY",
            "OTHER STUFF", "CHARITY"]
expenses.sort()


class ExcelTemplateCreate:

    def create_sheets(self):
        wb = Workbook()
        ws0 = wb.create_sheet("ANALISYS", 0)
        ws1 = wb.create_sheet(months[0], 1)
        ws2 = wb.create_sheet(months[1], 2)
        ws3 = wb.create_sheet(months[2], 3)
        ws4 = wb.create_sheet(months[3], 4)
        ws5 = wb.create_sheet(months[4], 5)
        ws6 = wb.create_sheet(months[5], 6)
        ws7 = wb.create_sheet(months[6], 7)
        ws8 = wb.create_sheet(months[7], 8)
        ws9 = wb.create_sheet(months[8], 9)
        ws10 = wb.create_sheet(months[9], 10)
        ws11 = wb.create_sheet(months[10], 11)
        ws12 = wb.create_sheet(months[11], 12)
        # print(wb.sheetnames)
        wb.save("budget.xlsx")

    def create_templates_for_analisys(self):
        wb = load_workbook("budget.xlsx")
        ws = wb['ANALISYS']

        ws['C2'] = wb.sheetnames[1]
        ws['D2'] = wb.sheetnames[2]
        ws['E2'] = wb.sheetnames[3]
        ws['F2'] = wb.sheetnames[4]
        ws['G2'] = wb.sheetnames[5]
        ws['H2'] = wb.sheetnames[6]
        ws['I2'] = wb.sheetnames[7]
        ws['J2'] = wb.sheetnames[8]
        ws['K2'] = wb.sheetnames[9]
        ws['L2'] = wb.sheetnames[10]
        ws['M2'] = wb.sheetnames[11]
        ws['N2'] = wb.sheetnames[12]

        ws['B3'] = expenses[0]
        ws['B4'] = expenses[1]
        ws['B5'] = expenses[2]
        ws['B6'] = expenses[3]
        ws['B7'] = expenses[4]
        ws['B8'] = expenses[5]
        ws['B9'] = expenses[6]
        ws['B10'] = expenses[7]
        ws['B11'] = expenses[8]
        ws['B12'] = expenses[9]
        ws['B13'] = expenses[10]
        ws['B14'] = expenses[11]
        ws['B15'] = expenses[12]
        ws['B16'] = expenses[13]
        ws['B17'] = expenses[14]
        ws['B18'] = expenses[15]
        ws['B19'] = expenses[16]
        ws['B20'] = expenses[17]
        ws['B21'] = expenses[18]

        ws.merge_cells('C1:N1')
        ws['C1'] = datetime.date.today().year
        wb.save("budget.xlsx")

    def create_templates_for_months(self):
        wb = load_workbook("budget.xlsx")
        ws1 = wb["JANUARY"]

        numbers = [x for x in range(32)]
        newlist = [int(x) for x in numbers if 0 < x < 32]
        newlist.insert(0, None)
        newlist.insert(0, None)

        for row in range(2):
            ws1.append(newlist)
            break

        ws1['C1'] = datetime.date.today().year
        ws1.merge_cells('C1:AG1')

        ws1['B3'] = expenses[0]
        ws1['B4'] = expenses[1]
        ws1['B5'] = expenses[2]
        ws1['B6'] = expenses[3]
        ws1['B7'] = expenses[4]
        ws1['B8'] = expenses[5]
        ws1['B9'] = expenses[6]
        ws1['B10'] = expenses[7]
        ws1['B11'] = expenses[8]
        ws1['B12'] = expenses[9]
        ws1['B13'] = expenses[10]
        ws1['B14'] = expenses[11]
        ws1['B15'] = expenses[12]
        ws1['B16'] = expenses[13]
        ws1['B17'] = expenses[14]
        ws1['B18'] = expenses[15]
        ws1['B19'] = expenses[16]
        ws1['B20'] = expenses[17]
        ws1['B21'] = expenses[18]
        # 31 days
        ws3 = wb["MARCH"]
        ws5 = wb["MAY"]
        ws7 = wb["JULY"]
        ws8 = wb["AUGUST"]
        ws10 = wb["OCTOBER"]
        ws12 = wb["DECEMBER"]
        # 30 days
        ws4 = wb["APRIL"]
        ws6 = wb["JUNE"]
        ws9 = wb["SEPTEMBER"]
        ws11 = wb["NOVEMBER"]
        # 29 days
        ws2 = wb["FEBRUARY"]
        maxr = ws1.max_row
        maxc = ws1.max_column
        for r in range(1, maxr + 1):
            for c in range(1, maxc + 1):
                ws2.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value
                ws3.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value
                ws4.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value
                ws5.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value
                ws6.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value
                ws7.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value
                ws8.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value
                ws9.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value
                ws10.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value
                ws11.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value
                ws12.cell(row=r, column=c).value = ws1.cell(row=r, column=c).value

        # TODO: create new file with copying sheets with 31 days in year
        # TODO: create months with 30 days and FEBRUARY
        # ws.delete_cols(30,1) #30 kolumna remove
        ws1.delete_rows(22,10)
        ws2.delete_rows(22, 10)
        wb.save("budget.xlsx")

'''
    def iteration(self):
        wb = load_workbook("budget.xlsx")
        ws = wb["JANUARY"]

        for x in expenses:
            print(x)

        choose_expense = input('Choose your expense: 1-19: ')
        choose_day = input('Choose your expense: 1-31: ')
        x = input('podaj kwote: ')

        if choose_expense == '1' and choose_day == '1':
            ws['C3'] = float(x)

        wb.save("budget.xlsx")

'''
'''
        for cell in ws["B"]:
            print(cell.value)


        
        ws['C2'] = 1
        ws['D2'] = 2
        ws['E2'] = 3
        ws['F2'] = 3
        ws['G2'] = 3
        ws['H2'] = 3
        ws['I2'] = 3
        ws['J2'] = 3
        ws['K2'] = 3
        ws['L2'] = 3
        ws['M2'] = 3
        ws['N2'] = 3
        ws['O2'] = 3
        ws['P2'] = 3
        ws['Q2'] = 3
        ws['R2'] = 3
        ws['S2'] = 3
        ws['T2'] = 3
        ws['U2'] = 3
        ws['V2'] = 3
        ws['W2'] = 3
        ws['X2'] = 3
        ws['Y2'] = 3
'''
# ws['A'].alignment = Alignment(horizontal="center")
# wb.save("budget.xlsx")


t = ExcelTemplateCreate()
# t.create_sheets()
# t.create_templates_for_analisys()
t.create_templates_for_months()
# t.iteration()
'''
        for row in range(1):
            ws1.append(range(0, 13))

        for cell in ws1['C']:
            cell.number_format = "DD/MM/YYYY"
'''
