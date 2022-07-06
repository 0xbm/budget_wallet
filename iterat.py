from natsort import os_sorted
from openpyxl import load_workbook

expenses = ["1.AGD RTV", "2.ANIMALS", "3.BEAUTY", "4.BILLS", "5.CAR", "6.CHARITY", "7.CLOTHES", "8.EATING OUT",
            "9.ENTERTAINMENT", "10.EVENTS", "11.GIFTS", "12.HOLIDAYS", "13.HOME SHOPPING", "14.INSURANCE",
            "15.MEDICINE", "16.OTHER STUFF", "17.REPAIRS", "18.SHOPPING", "19.SPORT", "q=QUIT"]

print(os_sorted(expenses))

def month():
    while (choose := input("Please select expensive month or q for quit: ")) != "q":

        if choose == "q":
            break
        if choose == "1":
            january()
        if choose == "2":
            february()


def january():
    wb = load_workbook("budget.xlsx")
    ws = wb["JANUARY"]

    for x in expenses:
        print(x)

    while (choose_expense := input("Choose your expense: 1-19 or q for quit: ")) != "q":
        choose_day = input("Choose your day: 1-31: ")
        x = input("podaj kwote: ")

        if choose_expense == "1" and choose_day == "1":
            print("You choose " + ws["B3"].value + " and day: " + choose_day)
            ws["C3"] = float(x)
        if choose_expense == "1" and choose_day == "2":
            print("You choose " + ws["B3"].value + " and day: " + choose_day)
            ws["D3"] = float(x)

    wb.save("budget.xlsx")

def february():
    wb = load_workbook("budget.xlsx")
    ws = wb["FEBRUARY"]

    for x in expenses:
        print(x)
    """
    while True:
        choose_expense = input("Choose your expense: 1-19: ")
        if choose_expense == "q":
            break
        choose_day = input("Choose your day: 1-31: ")
        x = input("podaj kwote: ")
    """

    while (choose_expense := input("Choose your expense: 1-19 or q for quit: ")) != "q":
        choose_day = input("Choose your day: 1-31: ")
        x = input("podaj kwote: ")

        if choose_expense == "1" and choose_day == "1":
            print("You choose " + ws["B3"].value + " and day: " + choose_day)
            ws["C3"] = float(x)
        if choose_expense == "1" and choose_day == "2":
            print("You choose " + ws["B3"].value + " and day: " + choose_day)
            ws["D3"] = float(x)
    wb.save("budget.xlsx")


month()
